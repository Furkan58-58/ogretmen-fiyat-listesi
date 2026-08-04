from __future__ import annotations

import json
import os
import re
from datetime import datetime
from html import escape
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import BaseDocTemplate, Frame, PageBreak, PageTemplate, Paragraph, Spacer, Table, TableStyle

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(os.environ.get("PRICE_WORKBOOK", ROOT / "output" / "Fiyat-Listesi-Yonetim.xlsx"))
DOCS = ROOT / "docs"
DATA = DOCS / "data" / "products.json"
PDF = DOCS / "downloads" / "Fiyat-Listesi.pdf"


def clean(value):
    return "" if value is None else str(value).strip()


def money_value(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_grade(value):
    text = clean(value)
    return "Maarif TYT" if text.casefold() == "maarif tyt" else text


def barcode_list(*values):
    result = []
    for value in values:
        for barcode in re.split(r"[,;\n]+", clean(value)):
            barcode = barcode.strip().lstrip("'")
            if barcode and barcode not in result:
                result.append(barcode)
    return result


def is_yes(value):
    return clean(value).casefold() in {"evet", "e", "yes", "1", "x", "aktif"}


def slug(value):
    table = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    return re.sub(r"[^a-z0-9]+", "-", clean(value).translate(table).casefold()).strip("-")


def anchor(text):
    base = re.sub(r"[^a-z0-9]+", "-", text.casefold().replace("ı", "i").replace("ş", "s").replace("ğ", "g").replace("ü", "u").replace("ö", "o").replace("ç", "c")).strip("-")
    return f"ders-{base}"


def load_data():
    wb = load_workbook(WORKBOOK, data_only=False)
    ws = wb["Ürünler"]
    rates = {}
    if "İndirimler" in wb.sheetnames:
        for publisher, grade, single_rate, bulk_rate in wb["İndirimler"].iter_rows(min_row=2, values_only=True):
            rates[(clean(publisher), normalize_grade(grade))] = (money_value(single_rate), money_value(bulk_rate))
    theme = {"title": "2026-2027 Öğretmen Fiyat Listesi", "subtitle": "Dersi seçin, sınıflara ayrılmış kitapları ve fiyatları görüntüleyin.", "mainColor": "#173F5F", "accentColor": "#F2B134"}
    if "Tasarım" in wb.sheetnames:
        settings = {clean(k): clean(v) for k, v in wb["Tasarım"].iter_rows(min_row=2, max_col=2, values_only=True) if k}
        theme.update({"title": settings.get("Başlık") or theme["title"], "subtitle": settings.get("Alt Başlık") or theme["subtitle"], "mainColor": settings.get("Ana Renk") or theme["mainColor"], "accentColor": settings.get("Vurgu Rengi") or theme["accentColor"]})
    grade_levels = {
        "5. Sınıf": "Ortaokul", "6. Sınıf": "Ortaokul", "7. Sınıf": "Ortaokul", "8. Sınıf": "Ortaokul",
        "9. Sınıf": "Lise", "10. Sınıf": "Lise", "11. Sınıf": "Lise", "12. Sınıf": "Lise", "Maarif TYT": "Lise",
    }
    bookstores = [("kitabevi-1", "1. Kitabevi"), ("kitabevi-2", "2. Kitabevi")]
    if "Ayarlar" in wb.sheetnames:
        settings_ws = wb["Ayarlar"]
        grade_levels = {normalize_grade(g): clean(level) for g, level in settings_ws.iter_rows(min_row=4, max_col=2, values_only=True) if g and level}
        configured = [(clean(code), clean(name)) for code, name in settings_ws.iter_rows(min_row=4, min_col=4, max_col=5, values_only=True) if code and name]
        if configured:
            bookstores = configured
    bookstore_names = dict(bookstores)
    first_bookstore = bookstores[0][0]
    publisher_bookstores = {}
    if "Ayarlar" in wb.sheetnames:
        settings_ws = wb["Ayarlar"]
        for publisher, bookstore_code in settings_ws.iter_rows(min_row=4, min_col=7, max_col=8, values_only=True):
            publisher, bookstore_code = clean(publisher), clean(bookstore_code)
            if publisher and bookstore_code:
                publisher_bookstores[publisher] = bookstore_code

    page_rows = []
    if "Sayfa Ayarları" in wb.sheetnames:
        page_ws = wb["Sayfa Ayarları"]
        page_headers = {clean(cell.value): index for index, cell in enumerate(page_ws[1])}
        for row in page_ws.iter_rows(min_row=2, values_only=True):
            def page_value(name):
                index = page_headers.get(name)
                return row[index] if index is not None and index < len(row) else None
            code, level = clean(page_value("Kitabevi Kodu")), clean(page_value("Kademe"))
            if not code or not level or not is_yes(page_value("Aktif")):
                continue
            page_rows.append({
                "bookstoreCode": code, "bookstoreName": bookstore_names.get(code, code), "level": level,
                "title": clean(page_value("Sayfa Başlığı")) or f"{bookstore_names.get(code, code)} {level}",
                "description": clean(page_value("Açıklama")) or f"{level} kitapları ve güncel fiyat listesi",
                "mainColor": clean(page_value("Ana Renk")) or theme["mainColor"],
                "accentColor": clean(page_value("Vurgu Rengi")) or theme["accentColor"],
                "logo": clean(page_value("Logo Linki")),
                "relatedKey": clean(page_value("Gösterilecek Diğer Liste")),
                "order": money_value(page_value("Ana Sayfa Sırası")) or 999,
            })
    if not page_rows:
        page_rows = [{"bookstoreCode": code, "bookstoreName": name, "level": level, "title": f"{name} {level}",
                      "description": f"{level} kitapları ve güncel fiyat listesi", "mainColor": theme["mainColor"],
                      "accentColor": theme["accentColor"], "order": order}
                     for order, (code, name, level) in enumerate([(bookstores[0][0], bookstores[0][1], "Ortaokul"), (bookstores[0][0], bookstores[0][1], "Lise")], 1)]

    contacts = []
    if "İletişim" in wb.sheetnames:
        contact_ws = wb["İletişim"]
        contact_headers = {clean(cell.value): index for index, cell in enumerate(contact_ws[1])}
        for row in contact_ws.iter_rows(min_row=2, values_only=True):
            def contact_value(name):
                index = contact_headers.get(name)
                return row[index] if index is not None and index < len(row) else None
            code, level, name, phone = map(clean, [contact_value("Kitabevi Kodu"), contact_value("Kademe"), contact_value("Ad Soyad"), contact_value("Telefon")])
            if code and level and (name or phone):
                contacts.append({"bookstoreCode": code, "level": level, "name": name, "phone": phone,
                                 "whatsapp": is_yes(contact_value("WhatsApp"))})

    headers = {clean(cell.value): index for index, cell in enumerate(ws[1])}
    def get(row, name):
        index = headers.get(name)
        return row[index] if index is not None and index < len(row) else None
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        publisher = clean(get(row, "Yayın"))
        grade = normalize_grade(get(row, "Sınıf"))
        course = clean(get(row, "Ders"))
        general = clean(get(row, "Genel Tür"))
        kind = clean(get(row, "Tür"))
        descriptor = general or kind
        book_name = " • ".join(x for x in [publisher, grade, course, descriptor] if x)
        price = money_value(get(row, "Fiyat"))
        single_rate, bulk_rate = rates.get((publisher, grade), (None, None))
        extra_barcodes = get(row, "Yeni Barkodlar")
        if extra_barcodes is None and len(row) > 10:
            extra_barcodes = row[10]
        barcodes = barcode_list(get(row, "Barkod"), extra_barcodes)
        bookstore_code = publisher_bookstores.get(publisher, first_bookstore)
        level = grade_levels.get(grade, "Lise")
        products.append({
            "barcode": barcodes[0] if barcodes else "", "barcodes": barcodes, "publisher": publisher, "grade": grade,
            "course": course, "category": general, "type": kind, "bookName": book_name,
            "price": price, "discountPrice": round(price * (1 - single_rate), 2) if price is not None and single_rate else None,
            "bulkPrice": round(price * (1 - bulk_rate), 2) if price is not None and bulk_rate else None,
            "singleRate": single_rate, "bulkRate": bulk_rate, "promo": clean(get(row, "Tanıtım Linki")),
            "bookstoreCode": bookstore_code, "level": level,
        })
    catalogs = []
    for page in sorted(page_rows, key=lambda x: x["order"]):
        page["contacts"] = [{"name": c["name"], "phone": c["phone"], "whatsapp": c["whatsapp"]} for c in contacts
                            if c["bookstoreCode"] == page["bookstoreCode"] and c["level"] == page["level"]]
        page["slug"] = f'{slug(page["bookstoreCode"])}-{slug(page["level"])}'
        page["url"] = f'?kitabevi={page["bookstoreCode"]}&kademe={slug(page["level"])}'
        page["pdf"] = f'downloads/Fiyat-Listesi-{page["slug"]}.pdf'
        page["count"] = sum(1 for p in products if p["bookstoreCode"] == page["bookstoreCode"] and p["level"] == page["level"])
        catalogs.append(page)
    catalog_lookup = {f'{item["bookstoreCode"]}|{item["level"]}'.replace(" ", "").casefold(): item for item in catalogs}
    for page in catalogs:
        related_key = clean(page.pop("relatedKey", "")).replace(" ", "").casefold()
        target = catalog_lookup.get(related_key) if related_key and related_key != "yok" else None
        if target is not None and target is not page:
            page["relatedCatalog"] = {
                "title": target["title"], "url": target["url"], "level": target["level"],
                "bookstoreName": target["bookstoreName"], "mainColor": target["mainColor"],
                "accentColor": target["accentColor"],
            }
    return products, theme, catalogs


def register_fonts():
    candidates = [
        (Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / "arial.ttf", Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts" / "arialbd.ttf"),
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")),
    ]
    for regular, bold in candidates:
        if regular.exists() and bold.exists():
            pdfmetrics.registerFont(TTFont("PriceSans", str(regular)))
            pdfmetrics.registerFont(TTFont("PriceSans-Bold", str(bold)))
            return "PriceSans", "PriceSans-Bold"
    return "Helvetica", "Helvetica-Bold"


def fmt_price(value):
    if value is None:
        return "-"
    return f"₺{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def create_pdf(products, updated, theme, pdf_path=PDF):
    regular, bold = register_fonts()
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    main_color = colors.HexColor(theme["mainColor"]); accent_color = colors.HexColor(theme["accentColor"])
    title = ParagraphStyle("TitleTR", parent=styles["Title"], fontName=bold, fontSize=20, leading=24, textColor=main_color, alignment=TA_LEFT)
    meta = ParagraphStyle("Meta", parent=styles["Normal"], fontName=regular, fontSize=8.5, textColor=colors.HexColor("#5B6570"))
    section = ParagraphStyle("Section", parent=styles["Heading2"], fontName=bold, fontSize=12, leading=15, textColor=colors.white)
    cell = ParagraphStyle("Cell", parent=styles["Normal"], fontName=regular, fontSize=6.8, leading=8.4)
    cell_bold = ParagraphStyle("CellBold", parent=cell, fontName=bold)
    price_style = ParagraphStyle("Price", parent=cell_bold, alignment=TA_RIGHT, textColor=colors.HexColor("#0F7A4D"))
    menu_style = ParagraphStyle("Menu", parent=styles["Normal"], fontName=bold, fontSize=10, leading=12, alignment=TA_CENTER, textColor=colors.HexColor("#173F5F"))

    def header_footer(canvas, doc):
        canvas.saveState(); canvas.setFont(regular, 7.5); canvas.setFillColor(colors.HexColor("#6B7280"))
        canvas.drawString(15 * mm, 10 * mm, "Öğretmen Fiyat Listesi"); canvas.drawRightString(195 * mm, 10 * mm, f"Sayfa {doc.page}"); canvas.restoreState()

    frame = Frame(14 * mm, 15 * mm, 182 * mm, 267 * mm, leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    doc = BaseDocTemplate(str(pdf_path), pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    doc.addPageTemplates(PageTemplate(id="main", frames=frame, onPage=header_footer))
    grouped = {}
    for p in products:
        grouped.setdefault(p["course"] or "Diğer", {}).setdefault(p["grade"] or "Diğer", []).append(p)
    courses = sorted(grouped, key=str.casefold)
    story = [Paragraph(escape(theme["title"]), title), Spacer(1, 2 * mm), Paragraph(f"{len(products)} ürün • Son güncelleme: {updated}", meta), Spacer(1, 7 * mm), Paragraph("Dersler", title), Spacer(1, 3 * mm)]
    menu_rows = []
    for i in range(0, len(courses), 3):
        row = [Paragraph(f'<link href="#{anchor(c)}" color="#173F5F">{escape(c)}</link>', menu_style) for c in courses[i:i+3]]
        row += [""] * (3 - len(row)); menu_rows.append(row)
    if menu_rows:
        menu = Table(menu_rows, colWidths=[58 * mm] * 3, hAlign="LEFT")
        menu.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F7FAFC")), ("BOX", (0, 0), (-1, -1), 0.8, main_color), ("INNERGRID", (0, 0), (-1, -1), 0.8, colors.white), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("TOPPADDING", (0, 0), (-1, -1), 9), ("BOTTOMPADDING", (0, 0), (-1, -1), 9), ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5)]))
        story += [menu]
    else:
        story += [Paragraph("Bu bölümde henüz ürün bulunmuyor.", meta)]
    grade_order = {name: index for index, name in enumerate(["9. Sınıf", "10. Sınıf", "11. Sınıf", "12. Sınıf", "Maarif TYT"])}
    for course in courses:
        story += [PageBreak(), Paragraph(f'<a name="{anchor(course)}"/>{escape(course)}', title), Paragraph("Sınıflara göre kitap ve fiyat listesi", meta), Spacer(1, 5 * mm)]
        for grade, items in sorted(grouped[course].items(), key=lambda x: (grade_order.get(x[0], 99), x[0].casefold())):
            head = Table([[Paragraph(escape(grade), section)]], colWidths=[182 * mm])
            head.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), main_color), ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)])); story.append(head)
            rows = [[Paragraph("Kitabın Adı", cell_bold), Paragraph("Barkod", cell_bold), Paragraph("Fiyatlar", cell_bold), "", ""], ["", "", Paragraph("Fiyat", cell_bold), Paragraph("İnd", cell_bold), Paragraph("Toplu", cell_bold)]]
            for p in items:
                name = escape(p["bookName"])
                if p["promo"]:
                    name = f'<link href="{escape(p["promo"], quote=True)}" color="#173F5F"><u>{name}</u></link>'
                rows.append([Paragraph(name, cell), Paragraph(escape(p["barcode"]), cell), Paragraph(fmt_price(p["price"]), price_style), Paragraph(fmt_price(p["discountPrice"]), price_style), Paragraph(fmt_price(p["bulkPrice"]), price_style)])
            table = Table(rows, colWidths=[91 * mm, 34 * mm, 19 * mm, 19 * mm, 19 * mm], repeatRows=2)
            table.setStyle(TableStyle([("SPAN", (0, 0), (0, 1)), ("SPAN", (1, 0), (1, 1)), ("SPAN", (2, 0), (4, 0)), ("BACKGROUND", (0, 0), (-1, 1), accent_color), ("ALIGN", (2, 0), (4, 1), "CENTER"), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9DEE5")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3), ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")])]))
            story += [table, Spacer(1, 4 * mm)]
    doc.build(story)


def main():
    products, theme, catalogs = load_data(); updated = datetime.now().strftime("%d.%m.%Y %H:%M")
    DATA.parent.mkdir(parents=True, exist_ok=True)
    DATA.write_text(json.dumps({"updated": updated, "count": len(products), "theme": theme, "catalogs": catalogs, "products": products}, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    create_pdf(products, updated, theme)
    for catalog in catalogs:
        catalog_products = [p for p in products if p["bookstoreCode"] == catalog["bookstoreCode"] and p["level"] == catalog["level"]]
        catalog_theme = {"title": catalog["title"], "subtitle": catalog["description"], "mainColor": catalog["mainColor"], "accentColor": catalog["accentColor"]}
        create_pdf(catalog_products, updated, catalog_theme, DOCS / catalog["pdf"])
    print(f"Generated {len(products)} products in {len(catalogs)} catalogs")


if __name__ == "__main__": main()
